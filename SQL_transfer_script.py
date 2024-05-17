import os
import pymysql
from openpyxl import load_workbook

# Database connection
database = pymysql.connect(
    host="localhost",
    user="root",
    passwd="12345678",
    db="first_sikorsky"
)
cursor = database.cursor()

# Create table queries
doc_type_table = """
CREATE TABLE IF NOT EXISTS doc_type_table (
    doc_type_id INT AUTO_INCREMENT PRIMARY KEY,
    doc_type VARCHAR(255) NOT NULL UNIQUE
)
"""
aircraft_model_table = """
CREATE TABLE IF NOT EXISTS aircraft_model_table (
    model_id INT AUTO_INCREMENT PRIMARY KEY,
    model_name VARCHAR(255) NOT NULL UNIQUE
)
"""
item_table = """
CREATE TABLE IF NOT EXISTS item_table (
    item_number_id INT AUTO_INCREMENT PRIMARY KEY,
    item_number_value VARCHAR(255) NOT NULL UNIQUE
)
"""

document_table = """
CREATE TABLE IF NOT EXISTS documents (
    Unique_ID INT AUTO_INCREMENT PRIMARY KEY,
    doc_type VARCHAR(255),
    Item_number VARCHAR(255),
    index_1 VARCHAR(255),
    index_2 VARCHAR(255),
    type VARCHAR(255),
    reference VARCHAR(255),
    see_it VARCHAR(255),
    Aircraft_Model VARCHAR(255),
    Description TEXT,
    names VARCHAR(255),
    date VARCHAR(20),
    comment TEXT,
    FOREIGN KEY (doc_type) REFERENCES doc_type_table(doc_type),
    FOREIGN KEY (Item_number) REFERENCES item_table(item_number_value),
    FOREIGN KEY (Aircraft_Model) REFERENCES aircraft_model_table(model_name)
)
"""
document_picture_table = """
CREATE TABLE IF NOT EXISTS document_picture (
    Unique_ID INT AUTO_INCREMENT PRIMARY KEY,
    doc_type VARCHAR(255),
    Item_number VARCHAR(255),
    index_1 VARCHAR(255),
    index_2 VARCHAR(255),
    type VARCHAR(255),
    reference VARCHAR(255),
    see_it VARCHAR(255),
    Aircraft_Model VARCHAR(255),
    Description TEXT,
    names VARCHAR(255),
    date VARCHAR(20),
    comment TEXT,
    FOREIGN KEY (doc_type) REFERENCES doc_type_table(doc_type),
    FOREIGN KEY (Item_number) REFERENCES item_table(item_number_value),
    FOREIGN KEY (Aircraft_Model) REFERENCES aircraft_model_table(model_name)
)
"""
hardware_table = """
CREATE TABLE IF NOT EXISTS hardware (
    Unique_ID INT AUTO_INCREMENT PRIMARY KEY,
    doc_type VARCHAR(255),
    Item_number VARCHAR(255),
    index_1 VARCHAR(255),
    index_2 VARCHAR(255),
    type VARCHAR(255),
    reference VARCHAR(255),
    see_it VARCHAR(255),
    Aircraft_Model VARCHAR(255),
    Description TEXT,
    names VARCHAR(255),
    date VARCHAR(20),
    comment TEXT,
    FOREIGN KEY (doc_type) REFERENCES doc_type_table(doc_type),
    FOREIGN KEY (Item_number) REFERENCES item_table(item_number_value),
    FOREIGN KEY (Aircraft_Model) REFERENCES aircraft_model_table(model_name)
)
"""
video_table = """
CREATE TABLE IF NOT EXISTS video (
    Unique_ID INT AUTO_INCREMENT PRIMARY KEY,
    doc_type VARCHAR(255),
    Item_number VARCHAR(255),
    index_1 VARCHAR(255),
    index_2 VARCHAR(255),
    type VARCHAR(255),
    reference VARCHAR(255),
    see_it VARCHAR(255),
    Aircraft_Model VARCHAR(255),
    Description TEXT,
    names VARCHAR(255),
    date VARCHAR(20),
    comment TEXT,
    FOREIGN KEY (doc_type) REFERENCES doc_type_table(doc_type),
    FOREIGN KEY (Item_number) REFERENCES item_table(item_number_value),
    FOREIGN KEY (Aircraft_Model) REFERENCES aircraft_model_table(model_name)
)
"""
picture_table = """
CREATE TABLE IF NOT EXISTS picture (
    Unique_ID INT AUTO_INCREMENT PRIMARY KEY,
    doc_type VARCHAR(255),
    Item_number VARCHAR(255),
    index_1 VARCHAR(255),
    index_2 VARCHAR(255),
    type VARCHAR(255),
    reference VARCHAR(255),
    see_it VARCHAR(255),
    Aircraft_Model VARCHAR(255),
    Description TEXT,
    names VARCHAR(255),
    date VARCHAR(20),
    comment TEXT,
    FOREIGN KEY (doc_type) REFERENCES doc_type_table(doc_type),
    FOREIGN KEY (Item_number) REFERENCES item_table(item_number_value),
    FOREIGN KEY (Aircraft_Model) REFERENCES aircraft_model_table(model_name)
)
"""
publication_table = """
CREATE TABLE IF NOT EXISTS publication (
    Unique_ID INT AUTO_INCREMENT PRIMARY KEY,
    doc_type VARCHAR(255),
    Item_number VARCHAR(255),
    index_1 VARCHAR(255),
    index_2 VARCHAR(255),
    type VARCHAR(255),
    reference VARCHAR(255),
    see_it VARCHAR(255),
    Aircraft_Model VARCHAR(255),
    Description TEXT,
    names VARCHAR(255),
    date VARCHAR(20),
    comment TEXT,
    FOREIGN KEY (doc_type) REFERENCES doc_type_table(doc_type),
    FOREIGN KEY (Item_number) REFERENCES item_table(item_number_value),
    FOREIGN KEY (Aircraft_Model) REFERENCES aircraft_model_table(model_name)
)
"""

# Execute the table creation queries
cursor.execute(doc_type_table)
cursor.execute(aircraft_model_table)
cursor.execute(item_table)
cursor.execute(document_table)
cursor.execute(document_picture_table)
cursor.execute(hardware_table)
cursor.execute(video_table)
cursor.execute(picture_table)
cursor.execute(publication_table)

# Path to the directory containing Excel files
excel_folders_path = r"/Users/joaquincamaran/Documents/Sikorsky(SAMPLE)/sample_excel_data"

# Valid doc types
valid_doc_types = {"Calendar", "Report", "Sikorsky News", "Notebook", "Manual", "File"}
valid_hardware_types = {"Model"}
valid_document_picture_types = {"Flat Document"}
valid_video_types = {"16mm Film", "8mm Film"}
valid_picture_types = {"Negative", "Photograph"}
valid_publication_types = {"Handout", "Book"}

# Functions to check if entries already exist in their respective tables
def model_exists(model_name):
    check_query = "SELECT model_id FROM aircraft_model_table WHERE model_name = %s"
    cursor.execute(check_query, (model_name,))
    return cursor.fetchone() is not None

def doc_type_exists(doc_type):
    check_query = "SELECT doc_type_id FROM doc_type_table WHERE doc_type = %s"
    cursor.execute(check_query, (doc_type,))
    return cursor.fetchone() is not None

def item_number_exists(item_number_value):
    check_query = "SELECT item_number_id FROM item_table WHERE item_number_value = %s"
    cursor.execute(check_query, (item_number_value,))
    return cursor.fetchone() is not None

# Iterate through all files in the directory
for file_name in os.listdir(excel_folders_path):
    if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
        file_path = os.path.join(excel_folders_path, file_name)
        print(f'Processing file: {file_path}')
        
        try:
            # Read the Excel file
            workbook = load_workbook(file_path)
        except Exception as e:
            print(f"Error loading workbook {file_path}: {e}")
            continue

        # Iterate through all sheets
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            unique_models = set()
            unique_doc_types = set()
            unique_item_numbers = set()

            # Iterate through rows and collect data
            for row_idx in range(2, sheet.max_row + 1):  # Skip header row
                doc_type = sheet.cell(row=row_idx, column=1).value  # Read data from column A
                item_number = sheet.cell(row=row_idx, column=2).value  # Read data from column B
                aircraft_model = sheet.cell(row=row_idx, column=8).value  # Read data from column H

                if doc_type and doc_type not in unique_doc_types:
                    unique_doc_types.add(doc_type)
                if item_number and item_number not in unique_item_numbers:
                    unique_item_numbers.add(item_number)
                if aircraft_model and aircraft_model not in unique_models:
                    unique_models.add(aircraft_model)

            # Insert unique doc_types into the doc_type_table
            for doc_type in unique_doc_types:
                if not doc_type_exists(doc_type):
                    insert_doc_type_query = "INSERT INTO doc_type_table (doc_type) VALUES (%s)"
                    cursor.execute(insert_doc_type_query, (doc_type,))

            # Insert unique item numbers into the item_table
            for item_number in unique_item_numbers:
                if not item_number_exists(item_number):
                    insert_item_number_query = "INSERT INTO item_table (item_number_value) VALUES (%s)"
                    cursor.execute(insert_item_number_query, (item_number,))

            # Insert unique models into the aircraft_model_table
            for model in unique_models:
                if not model_exists(model):
                    insert_model_query = "INSERT INTO aircraft_model_table (model_name) VALUES (%s)"
                    cursor.execute(insert_model_query, (model,))

# Commit the transaction
database.commit()
print("Data inserted successfully!")

# Iterate through all files in the directory again for detailed data insertion
for file_name in os.listdir(excel_folders_path):
    if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
        file_path = os.path.join(excel_folders_path, file_name)
        print(f'Processing file: {file_path}')

        try:
            # Read the Excel file
            workbook = load_workbook(file_path)
            sheet = workbook['Sheet1']  # Adjust sheet name as needed
        except Exception as e:
            print(f"Error loading workbook {file_path}: {e}")
            continue

        # Iterate through rows and columns to read data
        for row_idx in range(2, sheet.max_row + 1):  # Skip header row
            doc_type = sheet.cell(row=row_idx, column=1).value  # Read data from column A
            item_number = sheet.cell(row=row_idx, column=2).value  # Read data from column B
            index_1 = sheet.cell(row=row_idx, column=3).value  # Read data from column C
            index_2 = sheet.cell(row=row_idx, column=4).value  # Read data from column D
            type_value = sheet.cell(row=row_idx, column=5).value  # Read data from column E
            reference = sheet.cell(row=row_idx, column=6).value  # Read data from column F
            see_it = sheet.cell(row=row_idx, column=7).value  # Read data from column G
            aircraft_model = sheet.cell(row=row_idx, column=8).value  # Read data from column H
            description = sheet.cell(row=row_idx, column=9).value  # Read data from column I
            names = sheet.cell(row=row_idx, column=10).value  # Read data from column J
            date = sheet.cell(row=row_idx, column=11).value  # Read data from column K
            comment = sheet.cell(row=row_idx, column=12).value  # Read data from column L

            if see_it:
                see_it_file_path = os.path.join(excel_folders_path, see_it)
            else:
                see_it_file_path = None
            
            insert_query = None
            data = (doc_type, item_number, index_1, index_2, type_value, reference, see_it_file_path, aircraft_model, description, names, date, comment)
            
            if doc_type in valid_doc_types:
                # Insert data into the documents table
                insert_query = """
                INSERT INTO documents (doc_type, Item_number, index_1, index_2, type, reference, see_it, Aircraft_Model, Description, names, date, comment)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
            elif doc_type in valid_document_picture_types:
                # Insert data into the document picture table
                insert_query = """
                INSERT INTO document_picture (doc_type, Item_number, index_1, index_2, type, reference, see_it, Aircraft_Model, Description, names, date, comment)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
            elif doc_type in valid_hardware_types:
                # Insert data into the hardware table
                insert_query = """
                INSERT INTO hardware (doc_type, Item_number, index_1, index_2, type, reference, see_it, Aircraft_Model, Description, names, date, comment)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
            elif doc_type in valid_video_types:
                # Insert data into the video table
                insert_query = """
                INSERT INTO video (doc_type, Item_number, index_1, index_2, type, reference, see_it, Aircraft_Model, Description, names, date, comment)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
            elif doc_type in valid_picture_types:
                # Insert data into the picture table
                insert_query = """
                INSERT INTO picture (doc_type, Item_number, index_1, index_2, type, reference, see_it, Aircraft_Model, Description, names, date, comment)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
            elif doc_type in valid_publication_types:
                # Insert data into the publication table
                insert_query = """
                INSERT INTO publication (doc_type, Item_number, index_1, index_2, type, reference, see_it, Aircraft_Model, Description, names, date, comment)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
            else:
                print(f"Unknown doc_type '{doc_type}' in file '{file_name}'")
                continue  # Skip this row if doc_type does not match any known categories
            
            if insert_query:
                try:
                    cursor.execute(insert_query, data)
                except Exception as e:
                    print(f"Error inserting data for row {row_idx} in file {file_name}: {e}")

# Commit the transaction
database.commit()
print("Data inserted successfully!")
