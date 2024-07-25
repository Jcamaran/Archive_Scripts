import requests
import json
import datetime
from openpyxl import load_workbook

# Constants for API for Monday.com
apiKey = "YOUR_API_KEY"
apiUrl = "https://api.monday.com/v2"
headers = {"Authorization": apiKey}


# Path to Excel file
EXCEL_PATH = r"YOUR_PATH_TO_EXCEL FOLDERS"



def load_excel(file_path):
    """Load the Excel workbook and return the workbook object."""
    return load_workbook(file_path, data_only=True)

def list_worksheets(workbook):
    """List all worksheet names in the workbook."""
    print("These are all the worksheets in the file: ")
    for index, sheetname in enumerate(workbook.sheetnames, start=1):
        print(f"{index}: {sheetname}")

def get_worksheet(workbook, sheet_name):
    """Get the specified worksheet from the workbook."""
    return workbook[sheet_name]

def get_all_rows(worksheet):
    """Get all rows from the specified worksheet."""
    return list(worksheet.rows)



def enter_data_from_excel(rows):
    for i, row in enumerate(rows[65:68], start=1):
        request_number = row[0].value
        work_type = row[1].value
        date_entered = row[2].value
        clnd_tracking = row[3].value
        request_source = row[4].value
        request_from = row[5].value
        subject = row[6].value
        model = row[7].value
        description = row[8].value
        due_date = row[9].value
        critical = row[10].value
        contact_email = row[11].value
        assigned_archivist = row[12].value
        assigned_date = row[13].value
        request_status = row[14].value
        date_completed = row[15].value
        tracking_info = row[16].value
        material_link = row[17].value
        presentation_date_1 = row[18].value
        presentation_audience = row[19].value
        presentation_date_2 = row[20].value

        

        print(f"Row Number {i}: ")
        print(f"Request Number: {request_number}")
        print(f"Description: {description}")


        if isinstance(due_date, datetime.datetime):
            formatted_due_date = due_date.strftime('%m-%d-%Y')
            print(f"Due Date: {formatted_due_date}")
        else:
            formatted_due_date = None
        
        if isinstance(date_entered, datetime.datetime):
            formatted_date_entered = date_entered.strftime('%m-%d-%Y')
            print(f"Due Date: {formatted_date_entered}")
        else:
            formatted_date_entered = None

        if isinstance(assigned_date, datetime.datetime):
            formatted_assigned_date = assigned_date.strftime('%m-%d-%Y')
            print(f"Due Date: {formatted_assigned_date}")
        else:
            formatted_assigned_date = None
        
        if isinstance(date_completed, datetime.datetime):
            formatted_date_completed = date_completed.strftime('%m-%d-%Y')
            print(f"Due Date: {formatted_date_completed}")
        else:
            formatted_date_completed = None

        if critical == "Y":
            formatted_critical = "YES"
        else:
            formatted_critical = "NO"
    

        # Create the item
        create_item_query = """
        mutation($board_id: ID!, $item_name: String!) {
          create_item(board_id: $board_id, item_name: $item_name) {
            id
          }
        }
        """
        create_item_variables = {
            "board_id": "6951881761",  # Ensure the board ID is a string
            "item_name": f"Request: {subject}"
        }
        create_item_data = {'query': create_item_query, 'variables': create_item_variables}

        try:
            response = requests.post(url=apiUrl, json=create_item_data, headers=headers)
            response_data = response.json()
            if 'errors' in response_data:
                print(f"Error creating item: {response_data['errors']}")
            else:
                item_id = response_data['data']['create_item']['id']
                print(f"Item created with ID: {item_id}")

                create_update_query = """
                    mutation($item_id: ID!, $body: String!) {
                    create_update(item_id: $item_id, body: $body) {
                            id
                          }
                        }
                        """
                update_body = f"""Item created with the following details:
                    Request From: {request_from}
                    Contact Email: {contact_email}

                    Request Number: {request_number}

                    Subject: {subject}
                    Description: {description}

                    Archivist Assigned: {assigned_archivist}
                    Date Entered: {formatted_date_entered}
                    Assigned Date: {formatted_assigned_date}
                    Date Completed: {formatted_date_completed}
                    Due Date: {formatted_due_date}

                    Work Type: {work_type}
                    Calendar Tracking: {clnd_tracking}
                    Request Source: {request_source}
                    Model: {model}
                    Critical: {formatted_critical}
                    Request Status: {request_status}
                    Tracking Info: {tracking_info}
                    Material Link: {material_link}
                    Presentation Date: {presentation_date_1}
                    Presentation Audience: {presentation_audience}"""


                create_update_variables = {
                    "item_id": item_id,
                    "body": update_body
                }
                create_update_data = {'query': create_update_query, 'variables': create_update_variables}

                try:
                    update_response = requests.post(url=apiUrl, json=create_update_data, headers=headers)
                    update_response_data = update_response.json()
                    if 'errors' in update_response_data:
                        print(f"Error creating update: {update_response_data['errors']}")
                    else:
                        print(f"Update created for item ID: {item_id}")
                except requests.exceptions.RequestException as e:
                            print(f"Request failed: {e}")

                # #Update column values: Thus Section for some reason is not working in updating Monday.com
                # update_column_values_query = """
                # mutation($board_id: ID!, $item_id: ID!, $column_values: JSON!) {
                #   change_multiple_column_values(board_id: $board_id, item_id: $item_id, column_values: $column_values) {
                #     id
                #   }
                # }
                # """
                # column_values = {
                #     "status": {"label": "Done"},
                #     "date4": {"date": formatted_due_date},
                #     "email": {"email": "itsmyemail@mailserver.com", "text": "itsmyemail@mailserver.com"}
                # }
                # update_column_values_variables = {
                #     "board_id": "",  # Ensure the board ID is a string
                #     "item_id": item_id,        # Ensure the item ID is correct
                #     "column_values": json.dumps(column_values)
                # }
                # update_column_values_data = {'query': update_column_values_query, 'variables': update_column_values_variables}

                # try:
                #     update_response = requests.post(url=apiUrl, json=update_column_values_data, headers=headers)
                #     update_response_data = update_response.json()
                #     if 'errors' in update_response_data:
                #         print(f"Error updating column values: {update_response_data['errors']}")
                #     else:
                #         print(f"Column values updated for item ID: {item_id}")
                # except requests.exceptions.RequestException as e:
                #     print(f"Request failed: {e}")

        except requests.exceptions.RequestException as e:
            print(f"Request failed: {e}")

def main():
    # Load the workbook and list all worksheets
    workbook = load_excel(EXCEL_PATH)
    list_worksheets(workbook)
    
    # Get the tracking worksheet and all its rows
    tracking_worksheet = get_worksheet(workbook, 'Tracking')
    all_rows = get_all_rows(tracking_worksheet)
    print(f"There are {len(all_rows)} rows of data in this excel sheet")

    # Print the first five entries
    # print_first_five_entries(all_rows)
    enter_data_from_excel(all_rows)

if __name__ == "__main__":
    main()
